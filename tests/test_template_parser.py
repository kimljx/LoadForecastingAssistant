from pathlib import Path

import openpyxl

from core.template_parser import build_workbook_info, extract_all_metric_records
from core.workbook_loader import load_workbook_auto
from core.models import SheetRole


def make_workbook(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "十五五 十六五 110-35明细表"
    headers = ["序号", "项目名称", "所在区县", "电压等级", "增加变电容量", "开工年月", "投产年月", "投产年"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c).value = h
    ws2 = wb.create_sheet("表1 测试市容载比计算")
    ws2["A3"] = "年   份"
    for idx, year in enumerate(["2024年（现状）", "2025年（现状）", "2026年", "2030年"], start=3):
        ws2.cell(3, idx).value = year
    ws2["A35"] = "十八、110千伏网供负荷"
    ws2["C35"] = 100
    ws2["D35"] = 110
    ws2["E35"] = 120
    ws2["F35"] = 150
    ws2["A40"] = "十九、110千伏变电容量"
    ws2["C40"] = 200
    ws2["D40"] = 210
    ws2["E40"] = 220
    ws2["F40"] = 260
    ws2["A41"] = "二十、110千伏容载比"
    ws2["C41"] = 2.0
    ws2["D41"] = 1.909
    ws2["E41"] = 1.833
    ws2["F41"] = 1.733
    wb.save(path)


def test_dynamic_project_library_and_year_detection(tmp_path):
    path = tmp_path / "模板.xlsx"
    make_workbook(path)
    loaded = load_workbook_auto(path)
    info = build_workbook_info(loaded)
    roles = {s.name: s.role for s in info.sheet_infos.values()}
    assert roles["十五五 十六五 110-35明细表"] == SheetRole.PROJECT_LIBRARY_110_35
    assert info.latest_actual_year == 2025
    records = extract_all_metric_records(loaded, info)
    assert any(r.area == "测试市" and r.metric == "网供负荷" and r.year == 2030 for r in records)
