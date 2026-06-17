from pathlib import Path

import openpyxl

from core.models import Adjustment, ForecastRules, ScenarioResult
from core.workbook_loader import load_workbook_auto
from core.workbook_writer import export_scenario_to_workbook


def test_export_adds_result_sheet_and_does_not_modify_original(tmp_path):
    src = tmp_path / "原始.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "表1 测试市容载比计算"
    ws["A1"] = "测试"
    ws["B2"] = 100
    wb.save(src)
    loaded = load_workbook_auto(src)
    scenario = ScenarioResult(
        name="测试方案",
        description="测试",
        success=True,
        score=0,
        adjustments=[Adjustment("负荷", "测试市", "110kV", "网供负荷", 2030, 100, 120, 20, 0.2, "测试写回", "高", "低", "表1 测试市容载比计算", "B2")],
    )
    out = tmp_path / "结果.xlsx"
    export_scenario_to_workbook(loaded, scenario, ForecastRules(), out)
    orig = openpyxl.load_workbook(src)
    assert orig["表1 测试市容载比计算"]["B2"].value == 100
    new = openpyxl.load_workbook(out)
    assert "预测结果表" in new.sheetnames
    assert new["表1 测试市容载比计算"]["B2"].value == 120
