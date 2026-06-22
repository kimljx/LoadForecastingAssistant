from pathlib import Path

import openpyxl

from core.models import ForecastRules, ForecastTarget
from core.rule_persistence import (
    build_rule_profile,
    build_workbook_signature,
    compare_rule_profile,
    forecast_rules_from_dict,
    load_rule_profile,
    normalize_workbook_name,
    save_rule_profile,
    validate_target_presets,
)
from core.template_parser import build_workbook_info, extract_all_metric_records
from core.workbook_loader import load_workbook_auto


def _make_simple_template(path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "表1 测试市容载比计算"
    for col, year in enumerate(["2025年（现状）", "2026年", "2030年"], start=3):
        ws.cell(3, col).value = year
    ws["A35"] = "十八、110千伏网供负荷"
    ws["C35"] = 100
    ws["D35"] = 110
    ws["E35"] = 150
    ws["A40"] = "十九、110千伏变电容量"
    ws["C40"] = 200
    ws["D40"] = 220
    ws["E40"] = 260
    wb.save(path)


def test_normalize_workbook_name_removes_common_suffixes():
    assert normalize_workbook_name("供需预测模板_20260501.xlsx") == "供需预测模板"
    assert normalize_workbook_name("供需预测模板_反推结果_20260501_120000.xlsx") == "供需预测模板"
    assert normalize_workbook_name("供需预测模板_最终版.xlsx") == "供需预测模板"


def test_rule_profile_roundtrip_and_validation(tmp_path):
    path = tmp_path / "供需预测模板_20260501.xlsx"
    _make_simple_template(path)
    loaded = load_workbook_auto(path)
    info = build_workbook_info(loaded)
    records = extract_all_metric_records(loaded, info)
    rules = ForecastRules(latest_actual_year=2025, forecast_start_year=2026, ratio_min=1.3, ratio_max=2.5)
    targets = [ForecastTarget("容载比", "测试市", "110kV", year=2030, target_value=1.85)]
    profile = build_rule_profile(path, info, records, rules, targets)
    out = tmp_path / "供需预测模板_规则.yaml"
    save_rule_profile(out, profile)
    loaded_profile = load_rule_profile(out)
    restored = forecast_rules_from_dict(loaded_profile["rules"])
    assert restored.ratio_min == 1.3
    sig = build_workbook_signature(info, records)
    match = compare_rule_profile(loaded_profile, path, sig)
    assert match.level == "ok"
    validations = validate_target_presets(loaded_profile, records)
    assert validations[0].status == "matched"
