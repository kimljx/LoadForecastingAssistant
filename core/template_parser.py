"""供需预测模板识别与数据抽取。"""
from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import dataclass
from typing import Iterable

from openpyxl.utils import get_column_letter

from .models import MetricRecord, SheetInfo, SheetRole, WorkbookInfo, YearColumn
from .workbook_loader import LoadedWorkbook

YEAR_RE = re.compile(r"(20\d{2}|19\d{2})\s*年")

# 容载比表关键指标映射：通过行标签识别。
METRIC_PATTERNS: list[tuple[str, str, str]] = [
    (r"调度口径负荷", "总计", "调度口径负荷"),
    (r"220.*网供负荷", "220kV", "网供负荷"),
    (r"110.*网供负荷", "110kV", "网供负荷"),
    (r"110.*变电容量$|110.*变电容量（不含用户专用站）", "110kV", "变电容量"),
    (r"110.*容载比", "110kV", "容载比"),
    (r"35.*网供负荷", "35kV", "网供负荷"),
    (r"35.*变电容量$|35.*变电容量（不含用户专用站）", "35kV", "变电容量"),
    (r"35.*容载比", "35kV", "容载比"),
    (r"10.*网供负荷", "10kV", "网供负荷"),
    (r"10.*变电容量需求", "10kV", "变电容量需求"),
    (r"配变平均负载率", "10kV", "配变平均负载率"),
    (r"同时率", "总计", "同时率"),
    (r"区外送.*受.*电", "自动", "区外送受电"),
]

PROJECT_LIBRARY_REQUIRED = ["项目", "区", "电压", "投产"]
PROJECT_LIBRARY_OPTIONAL = ["容量", "增加", "开工", "建设"]


def _cell_text(value) -> str:
    """是什么：把单元格内容安全转成文本。

    为什么：模板中空值和数字混杂，统一处理可降低识别异常。
    """
    return "" if value is None else str(value).strip()


def _normalize_text(value) -> str:
    """是什么：清理文本中的空白字符。

    为什么：中文模板中常有空格换行，规范化后更容易做关键词识别。
    """
    return re.sub(r"\s+", "", _cell_text(value))


def find_year_columns(ws, max_header_rows: int = 5) -> list[YearColumn]:
    """是什么：在前几行中动态识别年份列，不限制到 2030。

为什么：供需预测模板经常改名和新增年份，解析方法必须说明稳定识别原因。"""
    found: dict[int, YearColumn] = {}
    for row in range(1, min(ws.max_row, max_header_rows) + 1):
        for col in range(1, ws.max_column + 1):
            raw = _cell_text(ws.cell(row, col).value)
            m = YEAR_RE.search(raw)
            if not m:
                continue
            year = int(m.group(1))
            # 同一年可能在多个表头层级出现，优先保留靠后的、更接近数据的行。
            found[col] = YearColumn(year=year, column_index=col, header_row=row, raw_header=raw, is_actual="现状" in raw)
    return sorted(found.values(), key=lambda x: x.column_index)


def _detect_area_from_sheet_name(name: str) -> str | None:
    """是什么：从容载比 sheet 名中识别区域名称。

    为什么：地市和区县是后续分摊、目标设置、结果展示的核心维度。
    """
    text = name.strip()
    # 表1 内江市容载比计算 / 表1-1 东兴区容载比计算
    m = re.search(r"表\d+(?:-\d+)?\s*([^\s]+?)容载比计算", text)
    if m:
        return m.group(1).strip()
    return None


def _sheet_contains_keywords(ws, keywords: Iterable[str], scan_rows: int = 6) -> int:
    """是什么：扫描 sheet 前几行并统计关键词命中。

    为什么：项目库 sheet 名可能变化，需要通过字段特征辅助识别。
    """
    text = " ".join(
        _normalize_text(ws.cell(r, c).value)
        for r in range(1, min(ws.max_row, scan_rows) + 1)
        for c in range(1, min(ws.max_column, 30) + 1)
    )
    return sum(1 for k in keywords if k in text)


def detect_sheet_roles(loaded: LoadedWorkbook) -> dict[str, SheetInfo]:
    """是什么：识别工作表角色。

项目库不按固定名称识别，而是按字段和 sheet 名辅助识别。

为什么：供需预测模板经常改名和新增年份，解析方法必须说明稳定识别原因。"""
    result: dict[str, SheetInfo] = {}
    for ws in loaded.formula_wb.worksheets:
        name = ws.title
        year_cols = find_year_columns(ws)
        info = SheetInfo(name=name, year_columns=year_cols)
        normalized_name = _normalize_text(name)
        area = _detect_area_from_sheet_name(name)

        if "校核" in normalized_name:
            info.role = SheetRole.CHECK
            info.confidence = 0.9
        elif "容载比计算" in normalized_name and re.match(r"表\s*1\s", name):
            info.role = SheetRole.CITY_RATIO
            info.area_name = area
            info.confidence = 0.95
        elif "容载比计算" in normalized_name and "表1-" in normalized_name:
            info.role = SheetRole.COUNTY_RATIO
            info.area_name = area
            info.confidence = 0.95
        elif "直供负荷明细" in normalized_name:
            info.role = SheetRole.DIRECT_LOAD
            info.voltage_level = _guess_voltage(normalized_name)
            info.confidence = 0.9
        elif "变电容量明细" in normalized_name:
            info.role = SheetRole.SUBSTATION_CAPACITY
            info.voltage_level = _guess_voltage(normalized_name)
            info.confidence = 0.9
        elif "电源装机明细" in normalized_name:
            info.role = SheetRole.POWER_SOURCE
            info.voltage_level = _guess_voltage(normalized_name)
            info.confidence = 0.85
        elif "储能装机明细" in normalized_name:
            info.role = SheetRole.STORAGE
            info.voltage_level = _guess_voltage(normalized_name)
            info.confidence = 0.85
        else:
            required_score = _sheet_contains_keywords(ws, PROJECT_LIBRARY_REQUIRED, scan_rows=5)
            optional_score = _sheet_contains_keywords(ws, PROJECT_LIBRARY_OPTIONAL, scan_rows=5)
            name_score = sum(k in normalized_name for k in ["110", "35", "明细", "项目", "规划"])
            if required_score >= 3 and optional_score >= 1 and name_score >= 2:
                info.role = SheetRole.PROJECT_LIBRARY_110_35
                info.confidence = 0.9
                info.notes.append("按字段识别为110/35kV项目库，未依赖固定五年规划名称。")
            else:
                info.role = SheetRole.UNKNOWN
                info.confidence = 0.0

        result[name] = info
    return result


def _guess_voltage(text: str) -> str | None:
    """是什么：从文本中推断电压等级。

    为什么：模板不同 sheet/行标签写法不完全一致，需要统一成 110kV/35kV/10kV。
    """
    if "110" in text:
        return "110kV"
    if "35" in text:
        return "35kV"
    if "10" in text:
        return "10kV"
    return None


def build_workbook_info(loaded: LoadedWorkbook, latest_actual_year: int = 2025) -> WorkbookInfo:
    """是什么：构建工作簿业务结构信息。

    为什么：后续抽取指标、规则匹配和导出写回都依赖该结构。
    """
    infos = detect_sheet_roles(loaded)
    all_years = [yc.year for info in infos.values() for yc in info.year_columns]
    actual_candidates = [yc.year for info in infos.values() for yc in info.year_columns if yc.is_actual]
    latest_actual = max(actual_candidates) if actual_candidates else latest_actual_year
    return WorkbookInfo(
        path=loaded.original_path,
        repaired_path=loaded.working_path if loaded.repaired else None,
        sheet_infos=infos,
        latest_actual_year=latest_actual,
        forecast_start_year=latest_actual + 1,
    )


def _match_metric(row_text: str) -> tuple[str, str] | None:
    """是什么：把表格行标签匹配为标准指标。

    为什么：原表行名较长且带序号，工具需要映射成稳定业务指标。
    """
    text = _normalize_text(row_text)
    for pattern, voltage, metric in METRIC_PATTERNS:
        if re.search(pattern, text):
            # 区外送受电需要根据行文字判断电压等级。
            if voltage == "自动":
                voltage = _guess_voltage(text) or "总计"
            return voltage, metric
    return None


def _get_value_cell(loaded: LoadedWorkbook, sheet_name: str, row: int, col: int):
    """是什么：读取公式缓存值或普通值。

    为什么：抽取指标时需要数值结果，不能只看公式字符串。
    """
    if sheet_name in loaded.value_wb.sheetnames:
        return loaded.value_wb[sheet_name].cell(row, col).value
    return loaded.formula_wb[sheet_name].cell(row, col).value


def _to_float(value) -> float | None:
    """是什么：把 Excel 单元格值安全转为浮点数。

    为什么：业务计算只处理数值，空白、布尔、文本应自动忽略。
    """
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    try:
        return float(value)
    except Exception:
        return None


def extract_ratio_sheet_records(loaded: LoadedWorkbook, info: SheetInfo, latest_actual_year: int) -> list[MetricRecord]:
    """是什么：抽取地市/区县容载比表中的核心指标。

为什么：供需预测模板经常改名和新增年份，解析方法必须说明稳定识别原因。"""
    ws = loaded.formula_wb[info.name]
    records: list[MetricRecord] = []
    area = info.area_name or info.name
    year_cols = info.year_columns
    if not year_cols:
        return records
    for row in range(1, ws.max_row + 1):
        row_label = " ".join(_cell_text(ws.cell(row, c).value) for c in (1, 2))
        match = _match_metric(row_label)
        if not match:
            continue
        voltage, metric = match
        for yc in year_cols:
            value = _to_float(_get_value_cell(loaded, info.name, row, yc.column_index))
            cell = ws.cell(row, yc.column_index)
            records.append(
                MetricRecord(
                    area=area,
                    voltage_level=voltage,
                    metric=metric,
                    year=yc.year,
                    value=value,
                    source_sheet=info.name,
                    source_cell=cell.coordinate,
                    source_row=row,
                    source_col=yc.column_index,
                    is_formula=isinstance(cell.value, str) and cell.value.startswith("="),
                    is_actual=yc.is_actual or yc.year <= latest_actual_year,
                )
            )
    return records




def _city_area_name(workbook_info: WorkbookInfo) -> str:
    """是什么：优先使用地市容载比表识别出的区域名。

为什么：供需预测模板经常改名和新增年份，解析方法必须说明稳定识别原因。"""
    for info in workbook_info.sheet_infos.values():
        if info.role == SheetRole.CITY_RATIO and info.area_name:
            return info.area_name
    return "地市"


def _match_check_metric(row_text: str) -> tuple[str, str] | None:
    """是什么：识别校核表中的横向指标。

目前重点抽取同时率，后续可扩展为偏差校核、容载比总览等。

为什么：供需预测模板经常改名和新增年份，解析方法必须说明稳定识别原因。"""
    text = _normalize_text(row_text)
    if "同时率" not in text:
        return None
    if "110" in text:
        return "110kV", "同时率"
    if "35" in text:
        return "35kV", "同时率"
    if "10" in text:
        return "10kV", "同时率"
    return "总计", "同时率"


def extract_check_sheet_records(loaded: LoadedWorkbook, info: SheetInfo, workbook_info: WorkbookInfo) -> list[MetricRecord]:
    """是什么：抽取校核表中的同时率等关键校核指标。

同时率在实际反推中属于低优先级可调变量，但必须被展示和校核。

为什么：供需预测模板经常改名和新增年份，解析方法必须说明稳定识别原因。"""
    ws = loaded.formula_wb[info.name]
    area = _city_area_name(workbook_info)
    records: list[MetricRecord] = []
    year_cols = info.year_columns
    if not year_cols:
        return records
    for row in range(1, ws.max_row + 1):
        row_label = " ".join(_cell_text(ws.cell(row, c).value) for c in (1, 2))
        match = _match_check_metric(row_label)
        if not match:
            continue
        voltage, metric = match
        for yc in year_cols:
            value = _to_float(_get_value_cell(loaded, info.name, row, yc.column_index))
            cell = ws.cell(row, yc.column_index)
            records.append(
                MetricRecord(
                    area=area,
                    voltage_level=voltage,
                    metric=metric,
                    year=yc.year,
                    value=value,
                    source_sheet=info.name,
                    source_cell=cell.coordinate,
                    source_row=row,
                    source_col=yc.column_index,
                    is_formula=isinstance(cell.value, str) and cell.value.startswith("="),
                    is_actual=yc.is_actual or yc.year <= workbook_info.latest_actual_year,
                )
            )
    return records


def extract_all_metric_records(loaded: LoadedWorkbook, workbook_info: WorkbookInfo) -> list[MetricRecord]:
    """是什么：抽取当前模板中可用于预测和反推的核心指标。

为什么：供需预测模板经常改名和新增年份，解析方法必须说明稳定识别原因。"""
    records: list[MetricRecord] = []
    for info in workbook_info.sheet_infos.values():
        if info.role in {SheetRole.CITY_RATIO, SheetRole.COUNTY_RATIO}:
            records.extend(extract_ratio_sheet_records(loaded, info, workbook_info.latest_actual_year))
        elif info.role == SheetRole.CHECK:
            records.extend(extract_check_sheet_records(loaded, info, workbook_info))
    return records


def records_to_key_map(records: Iterable[MetricRecord]) -> dict[tuple[str, str, str, int], MetricRecord]:
    """是什么：把指标记录列表转为快速查找字典。

    为什么：求解器需要频繁定位某区域某年某指标。
    """
    return {(r.area, r.voltage_level, r.metric, r.year): r for r in records}


def list_areas(records: Iterable[MetricRecord]) -> list[str]:
    """是什么：列出模板中的区域选项。

    为什么：界面目标设置需要中文下拉列表。
    """
    return sorted({r.area for r in records})


def list_years(records: Iterable[MetricRecord]) -> list[int]:
    """是什么：列出模板中的年份集合。

    为什么：界面需要自动设置预测起止年和新增年份范围。
    """
    return sorted({r.year for r in records})


def list_voltage_levels(records: Iterable[MetricRecord]) -> list[str]:
    """是什么：列出模板中的电压等级选项。

    为什么：目标设置和规则校核都围绕电压等级展开。
    """
    order = {"总计": 0, "220kV": 1, "110kV": 2, "35kV": 3, "10kV": 4}
    return sorted({r.voltage_level for r in records}, key=lambda x: order.get(x, 99))
