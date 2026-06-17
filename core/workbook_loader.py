"""Excel 安全读取模块。

设计重点：
1. 不依赖 WPS/Excel COM；
2. xlsx 使用 openpyxl；
3. xls 使用 xlrd 读取并转换为临时 xlsx 分析；
4. 对 WPS/Excel 可能生成的非法筛选/排序 XML 做“副本修复”，不修改原文件。
"""
from __future__ import annotations

import re
import shutil
import tempfile
from dataclasses import dataclass
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import openpyxl
from openpyxl import Workbook

from .exceptions import UnsupportedFileTypeError, WorkbookLoadError


@dataclass(slots=True)
class LoadedWorkbook:
    """已经安全加载的工作簿。

    formula_wb: 保留公式的工作簿。
    value_wb: 使用缓存值读取的工作簿，用于获取公式结果。
    working_path: 实际被 openpyxl 读取的路径，可能是修复后的临时副本。
    original_path: 用户传入的原始路径，绝不修改。
    repaired: 是否生成过修复副本。
    source_format: xlsx 或 xls。
    """

    formula_wb: openpyxl.Workbook
    value_wb: openpyxl.Workbook
    working_path: Path
    original_path: Path
    repaired: bool
    source_format: str


_INVALID_ROW_ONLY_REF = re.compile(r'ref="\d+:\d+"')


def _repair_xlsx_to_temp(path: Path) -> Path:
    """修复 xlsx 中 openpyxl 无法读取的部分 XML。

    典型问题：WPS/Excel 生成的 autoFilter/sortState 使用了 row-only ref，
    例如 sortState ref="3:61"，openpyxl 会认为不是合法单元格区域。

    处理方式：只在临时副本中删除该 sortState 节点，不改动用户原文件。
    """
    tmp_dir = Path(tempfile.mkdtemp(prefix="lfa_xlsx_repair_"))
    repaired = tmp_dir / path.name
    try:
        with ZipFile(path, "r") as zin, ZipFile(repaired, "w", ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename.startswith("xl/worksheets/") and item.filename.endswith(".xml"):
                    text = data.decode("utf-8", errors="replace")
                    # 删除非法 row-only sortState，保留 autoFilter 本身。
                    text = re.sub(
                        r'<sortState\s+ref="\d+:\d+"[^>]*>.*?</sortState>',
                        "",
                        text,
                        flags=re.DOTALL,
                    )
                    text = re.sub(
                        r'<sortState\s+ref="\d+:\d+"[^>]*/>',
                        "",
                        text,
                        flags=re.DOTALL,
                    )
                    data = text.encode("utf-8")
                zout.writestr(item, data)
    except Exception as exc:  # pragma: no cover - 极端损坏文件
        raise WorkbookLoadError(f"修复 xlsx 副本失败：{exc}") from exc
    return repaired


def _load_xlsx(path: Path) -> LoadedWorkbook:
    repaired = False
    working_path = path
    try:
        formula_wb = openpyxl.load_workbook(path, data_only=False)
        value_wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        working_path = _repair_xlsx_to_temp(path)
        repaired = True
        try:
            formula_wb = openpyxl.load_workbook(working_path, data_only=False)
            value_wb = openpyxl.load_workbook(working_path, data_only=True)
        except Exception as exc:
            raise WorkbookLoadError(f"xlsx 读取失败，修复副本后仍无法读取：{exc}") from exc
    return LoadedWorkbook(
        formula_wb=formula_wb,
        value_wb=value_wb,
        working_path=working_path,
        original_path=path,
        repaired=repaired,
        source_format="xlsx",
    )


def _xls_to_temp_xlsx(path: Path) -> Path:
    """使用 xlrd 读取 xls 并转换为临时 xlsx。

    说明：xls 的公式、样式和合并单元格支持有限。第一版主要保证分析数据可读，
    导出仍统一为 xlsx。
    """
    try:
        import xlrd  # type: ignore
    except ImportError as exc:  # pragma: no cover
        raise WorkbookLoadError("读取 .xls 需要安装 xlrd==1.2.0") from exc

    try:
        book = xlrd.open_workbook(str(path), formatting_info=False)
    except Exception as exc:
        raise WorkbookLoadError(f"xls 读取失败：{exc}") from exc

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for sheet in book.sheets():
        ws = wb.create_sheet(sheet.name[:31])
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                ws.cell(r + 1, c + 1).value = sheet.cell_value(r, c)
    tmp_dir = Path(tempfile.mkdtemp(prefix="lfa_xls_convert_"))
    out = tmp_dir / f"{path.stem}_converted.xlsx"
    wb.save(out)
    return out


def _load_xls(path: Path) -> LoadedWorkbook:
    converted = _xls_to_temp_xlsx(path)
    formula_wb = openpyxl.load_workbook(converted, data_only=False)
    value_wb = openpyxl.load_workbook(converted, data_only=True)
    return LoadedWorkbook(
        formula_wb=formula_wb,
        value_wb=value_wb,
        working_path=converted,
        original_path=path,
        repaired=True,
        source_format="xls",
    )


def load_workbook_auto(path: str | Path) -> LoadedWorkbook:
    """根据扩展名自动读取工作簿。"""
    p = Path(path)
    if not p.exists():
        raise WorkbookLoadError(f"文件不存在：{p}")
    ext = p.suffix.lower()
    if ext == ".xlsx":
        return _load_xlsx(p)
    if ext == ".xls":
        return _load_xls(p)
    raise UnsupportedFileTypeError(f"暂不支持的文件类型：{ext}。仅支持 .xlsx / .xls")


def copy_workbook_for_export(loaded: LoadedWorkbook, output_path: str | Path) -> Path:
    """创建导出副本。

    xlsx 输入：优先复制可读取的 working_path；若发生过 XML 修复，则复制修复副本。
    xls 输入：复制转换后的 xlsx。
    """
    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(loaded.working_path, out)
    return out
