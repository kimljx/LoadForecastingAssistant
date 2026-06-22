"""Excel 导出与写回模块。"""
from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import Adjustment, ForecastRules, MetricRecord, ScenarioResult, WorkbookInfo
from .template_parser import YEAR_RE
from .workbook_loader import LoadedWorkbook, copy_workbook_for_export


def build_output_path(input_path: str | Path, output_dir: str | Path | None = None) -> Path:
    """是什么：生成默认导出文件名。

    为什么：所有结果必须另存为新文件，避免误改原始模板。
    """
    p = Path(input_path)
    out_dir = Path(output_dir) if output_dir else p.parent
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return out_dir / f"{p.stem}_反推结果_{stamp}.xlsx"


def _write_cell(ws, cell: str, value, overwrite_formula: bool) -> tuple[bool, str]:
    """是什么：向目标单元格写入数值并尊重公式覆盖开关。

    为什么：有些位置是公式，默认策略必须可控，避免破坏模板公式。
    """
    c = ws[cell]
    if isinstance(c.value, str) and c.value.startswith("=") and not overwrite_formula:
        return False, "目标单元格为公式，按规则未覆盖"
    c.value = value
    return True, "已写入"


def _copy_column_style(ws, from_col: int, to_col: int) -> None:
    """是什么：复制年份列样式到新增年份列。

    为什么：2031+ 新增列应尽量保持原模板外观一致。
    """
    ws.column_dimensions[get_column_letter(to_col)].width = ws.column_dimensions[get_column_letter(from_col)].width
    for row in range(1, ws.max_row + 1):
        src = ws.cell(row, from_col)
        dst = ws.cell(row, to_col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.alignment:
            dst.alignment = copy(src.alignment)


def _find_year_columns(ws) -> dict[int, int]:
    """是什么：定位工作表中的年份列。

    为什么：写回新增年份时必须知道已有年份列位置。
    """
    result: dict[int, int] = {}
    for row in range(1, min(ws.max_row, 5) + 1):
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row, col).value
            if val is None:
                continue
            m = YEAR_RE.search(str(val))
            if m:
                result[int(m.group(1))] = col
    return result


def _ensure_year_column(ws, year: int) -> int:
    """是什么：确保某年列存在。若不存在，在最后一个年份列后插入新列。

第一版仅复制样式和值写入，不尝试完整维护所有跨列公式。

为什么：写回副本关系到原始模板安全和新增年份扩展，必须说明写回策略原因。"""
    cols = _find_year_columns(ws)
    if year in cols:
        return cols[year]
    if not cols:
        raise ValueError(f"工作表 {ws.title} 未找到年份列，无法扩展 {year} 年")
    last_year = max(cols)
    last_col = cols[last_year]
    insert_at = last_col + 1
    ws.insert_cols(insert_at)
    _copy_column_style(ws, last_col, insert_at)
    # 找到上一年表头所在行。
    header_row = None
    for row in range(1, min(ws.max_row, 5) + 1):
        if ws.cell(row, last_col).value and YEAR_RE.search(str(ws.cell(row, last_col).value)):
            header_row = row
            break
    if header_row is None:
        header_row = 3
    ws.cell(header_row, insert_at).value = f"{year}年"
    return insert_at


def _resolve_write_cell(ws, adj: Adjustment) -> tuple[str | None, str]:
    """是什么：根据调整项定位写回单元格。

第二版优先使用 source_row + 年份列定位。这样当预测 2031、2035
等模板中不存在的年份时，可以自动扩展年份列并写入正确行，避免误写到
2030 等已有年份单元格。

为什么：写回副本关系到原始模板安全和新增年份扩展，必须说明写回策略原因。"""
    if adj.source_row:
        try:
            col = _ensure_year_column(ws, adj.year)
        except Exception as exc:
            return None, f"无法定位或扩展年份列：{exc}"
        return ws.cell(adj.source_row, col).coordinate, "按指标行和年份列定位"
    if adj.source_cell:
        return adj.source_cell, "按来源单元格定位"
    return None, "缺少来源行/来源单元格"


def _write_adjustments_to_original_sheets(
    wb: openpyxl.Workbook,
    scenario: ScenarioResult,
    rules: ForecastRules,
    overwrite_formula: bool,
) -> list[dict]:
    """是什么：把用户确认的调整项写回导出副本。

第二版改进：
- 对 2031+ 等模板不存在的年份，使用 source_row + 动态新增年份列写入；
- 对已有年份，也优先使用“来源行 + 年份列”定位，避免 source_cell 是旧年份单元格时写错；
- 所有跳过/写入动作完整记录到“预测结果表”。

为什么：写回副本关系到原始模板安全和新增年份扩展，必须说明写回策略原因。"""
    log: list[dict] = []
    for adj in scenario.adjustments:
        if not adj.write_back:
            log.append({
                "工作表": adj.source_sheet or "",
                "单元格": adj.source_cell or "",
                "年份": adj.year,
                "指标": adj.metric,
                "原值": adj.old_value,
                "写入值": adj.new_value,
                "状态": "未写回",
                "原因": adj.note or "该调整项仅作为方案提示",
            })
            continue
        if adj.year <= rules.latest_actual_year:
            log.append({"工作表": adj.source_sheet or "", "单元格": adj.source_cell or "", "年份": adj.year, "指标": adj.metric, "状态": "跳过", "原因": "现状年不可修改"})
            continue
        if not adj.source_sheet or adj.source_sheet not in wb.sheetnames:
            log.append({"工作表": adj.source_sheet or "", "单元格": adj.source_cell or "", "年份": adj.year, "指标": adj.metric, "状态": "未写回", "原因": "缺少来源工作表，通常是兜底变量或非模板指标"})
            continue
        ws = wb[adj.source_sheet]
        target_cell = ""
        try:
            if adj.source_row:
                target_col = _ensure_year_column(ws, adj.year)
                target_cell = ws.cell(adj.source_row, target_col).coordinate
            elif adj.source_cell:
                target_cell = adj.source_cell
            else:
                log.append({"工作表": adj.source_sheet, "单元格": "", "年份": adj.year, "指标": adj.metric, "状态": "未写回", "原因": "缺少来源行/单元格，无法定位写回位置"})
                continue
        except Exception as exc:
            log.append({"工作表": adj.source_sheet, "单元格": adj.source_cell or "", "年份": adj.year, "指标": adj.metric, "状态": "未写回", "原因": f"定位年份列失败：{exc}"})
            continue
        ok, msg = _write_cell(ws, target_cell, adj.new_value, overwrite_formula=overwrite_formula)
        log.append(
            {
                "工作表": adj.source_sheet,
                "单元格": target_cell,
                "年份": adj.year,
                "指标": adj.metric,
                "原值": adj.old_value,
                "写入值": adj.new_value,
                "状态": "已写回" if ok else "跳过",
                "原因": msg + (f"；{adj.note}" if adj.note else ""),
            }
        )
    return log

def _autosize(ws, min_width: int = 10, max_width: int = 38) -> None:
    # 合并单元格会产生 MergedCell，不能依赖 col[0].column_letter。
    """是什么：自动调整结果表列宽。

    为什么：中文结果说明较长，自动列宽能提升可读性。
    """
    for idx, col in enumerate(ws.columns, start=1):
        letter = get_column_letter(idx)
        max_len = 0
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))


def add_result_sheet(wb: openpyxl.Workbook, scenario: ScenarioResult, rules: ForecastRules, write_log: list[dict]) -> None:
    """是什么：新增并填充预测结果表。

    为什么：导出副本不仅要写值，还要解释目标达成、调整原因和写回清单。
    """
    name = "预测结果表"
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name, 0)
    title_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)

    row = 1
    ws.cell(row, 1).value = "负荷预测与容载比反推结果"
    ws.cell(row, 1).font = Font(size=16, bold=True, color="FFFFFF")
    ws.cell(row, 1).fill = title_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    row += 2

    basic = [
        ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("方案名称", scenario.name),
        ("方案说明", scenario.description),
        ("最新现状年", rules.latest_actual_year),
        ("预测起始年", rules.forecast_start_year),
        ("容载比硬边界", f"{rules.ratio_min}-{rules.ratio_max}"),
        ("同时率边界", f"{rules.coincidence_factor_min}-{rules.coincidence_factor_max}，单次调整不超过{rules.coincidence_factor_max_abs_change}"),
    ]
    for k, v in basic:
        ws.cell(row, 1).value = k
        ws.cell(row, 1).font = bold_font
        ws.cell(row, 2).value = v
        row += 1
    row += 1

    def section(title: str):
        """是什么：在结果表中写入分节标题。

        为什么：结果表内容较多，分节可以让业务人员快速阅读。
        """
        nonlocal row
        ws.cell(row, 1).value = title
        ws.cell(row, 1).font = bold_font
        ws.cell(row, 1).fill = section_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        row += 1

    def write_table(headers: list[str], rows: list[list]):
        """是什么：在结果表中写入标准二维表。

        为什么：统一写表逻辑可减少格式不一致和重复代码。
        """
        nonlocal row
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row, i)
            c.value = h
            c.font = bold_font
            c.fill = PatternFill("solid", fgColor="E2F0D9")
        row += 1
        for values in rows:
            for i, v in enumerate(values, start=1):
                ws.cell(row, i).value = v
            row += 1
        row += 1

    section("一、目标达成情况")
    target_rows = []
    for tr in scenario.target_results:
        t = tr.target
        target_rows.append([
            t.target_type,
            t.area,
            t.voltage_level,
            t.year or f"{t.period_start}-{t.period_end}",
            t.metric or "",
            t.target_value if t.target_value is not None else f"{t.min_value or ''}-{t.max_value or ''}",
            tr.actual_value,
            "达标" if tr.achieved else "未达标",
            tr.deviation,
            tr.message,
        ])
    write_table(["目标类型", "区域", "电压等级", "年份/阶段", "指标", "目标", "预测结果", "是否达标", "偏差", "说明"], target_rows)

    section("二、调整项明细")
    adj_rows = []
    for a in scenario.adjustments:
        adj_rows.append([
            a.object_type, a.area, a.voltage_level, a.metric, a.year, a.old_value, a.new_value, a.delta, a.delta_pct,
            a.reason, a.priority, a.risk_level, a.source_sheet or "", a.source_cell or "", a.note,
        ])
    write_table(["调整对象", "区域", "电压等级", "指标", "年份", "原值", "建议值", "变化量", "变化比例", "调整原因", "优先级", "风险", "来源表", "来源单元格", "备注"], adj_rows)

    section("三、预测结果明细")
    records = sorted(scenario.forecast_records, key=lambda r: (r.area, r.voltage_level, r.metric, r.year))
    pred_rows = [
        [r.area, r.voltage_level, r.metric, r.year, r.value, r.source_sheet, r.source_cell, "是" if r.is_actual else "否"]
        for r in records
    ]
    write_table(["区域", "电压等级", "指标", "年份", "预测/结果值", "来源表", "来源单元格", "是否现状年"], pred_rows)

    section("四、增长率结果")
    value_map = {(r.area, r.voltage_level, r.metric, r.year): r.value for r in records}
    growth_rows = []
    for r in records:
        if r.metric != "网供负荷" or r.year <= rules.latest_actual_year:
            continue
        prev = value_map.get((r.area, r.voltage_level, r.metric, r.year - 1))
        if prev in (None, 0) or r.value is None:
            continue
        g = r.value / prev - 1
        growth_rows.append([r.area, r.voltage_level, r.metric, r.year, prev, r.value, g, "达标" if rules.annual_growth_min <= g <= rules.annual_growth_max else "预警"])
    write_table(["区域", "电压等级", "指标", "年份", "上年值", "本年值", "单年增长率", "校核"], growth_rows)

    section("五、关键指标校核")
    check_rows = []
    for r in records:
        if r.metric not in {"容载比", "配变平均负载率", "同时率"} or r.year <= rules.latest_actual_year:
            continue
        if r.metric == "容载比":
            ok = rules.ratio_min <= (r.value or 0) <= rules.ratio_max
            limit = f"{rules.ratio_min}-{rules.ratio_max}"
        elif r.metric == "配变平均负载率":
            ok = rules.transformer_load_rate_min <= (r.value or 0) <= rules.transformer_load_rate_max
            limit = f"{rules.transformer_load_rate_min}-{rules.transformer_load_rate_max}，软目标{rules.transformer_load_rate_soft_target}"
        else:
            ok = rules.coincidence_factor_min <= (r.value or 0) <= rules.coincidence_factor_max
            limit = f"{rules.coincidence_factor_min}-{rules.coincidence_factor_max}"
        check_rows.append([r.area, r.voltage_level, r.metric, r.year, r.value, limit, "达标" if ok else "预警"])
    write_table(["区域", "电压等级", "指标", "年份", "结果值", "规则范围", "校核"], check_rows)

    section("六、写回单元格清单")
    log_rows = []
    for item in write_log:
        log_rows.append([
            item.get("工作表", ""), item.get("单元格", ""), item.get("年份", ""), item.get("指标", ""), item.get("原值", ""),
            item.get("写入值", ""), item.get("状态", ""), item.get("原因", ""),
        ])
    write_table(["工作表", "单元格", "年份", "指标", "原值", "写入值", "状态", "原因"], log_rows)

    if scenario.warnings:
        section("七、风险提示")
        for w in scenario.warnings:
            ws.cell(row, 1).value = w
            row += 1

    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    _autosize(ws)


def export_scenario_to_workbook(
    loaded: LoadedWorkbook,
    scenario: ScenarioResult,
    rules: ForecastRules,
    output_path: str | Path,
    overwrite_formula: bool = True,
) -> Path:
    """是什么：复制原工作簿，写回方案，并新增“预测结果表”。

原始传入文件绝不修改。

为什么：写回副本关系到原始模板安全和新增年份扩展，必须说明写回策略原因。"""
    out = copy_workbook_for_export(loaded, output_path)
    wb = openpyxl.load_workbook(out)
    write_log = _write_adjustments_to_original_sheets(wb, scenario, rules, overwrite_formula=overwrite_formula)
    add_result_sheet(wb, scenario, rules, write_log)
    wb.save(out)
    return out
